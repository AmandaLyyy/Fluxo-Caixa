"""
Company Fluxo de Caixa · Importador de Planilha
=============================================
Lê o arquivo fluxo_XX_XX_XX.xlsx e popula o banco MySQL cashflow_db.

Uso:
    pip install openpyxl mysql-connector-python python-dotenv
    python scripts/importar_planilha.py caminho/para/fluxo.xlsx
"""

import sys
import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

import openpyxl
import mysql.connector
from dotenv import load_dotenv

load_dotenv()

# ── Conexão ──────────────────────────────────────────────────────────────────
def conectar():
    return mysql.connector.connect(
        host     = os.getenv("DB_HOST", "localhost"),
        port     = int(os.getenv("DB_PORT", 3306)),
        user     = os.getenv("DB_USER", "root"),
        password = os.getenv("DB_PASSWORD", ""),
        database = os.getenv("DB_NAME", "cashflow_db"),
    )

# ── Helpers ───────────────────────────────────────────────────────────────────
def to_dec(v):
    """Converte qualquer valor numérico para Decimal seguro."""
    if v is None:
        return Decimal("0")
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")

def to_date(v):
    """Converte datetime/date/string para date."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None

def str_limpa(v):
    """String sem espaços extras."""
    return str(v).strip() if v is not None else ""

# ── Bancos ────────────────────────────────────────────────────────────────────
BANCOS_PLANILHA = [
    "BRADESCO", "SANTANDER", "ITAU", "BB", "Safra",
    "Caixa Corporate", "Bradesco HSBC", "ABC", "BOCOM",
    "SICREDI", "ASA", "BV",
]

ENDIVIDAMENTO_PLANILHA = {
    "Itaú": "ITAU", "Santander": "SANTANDER", "Safra": "Safra",
    "Bradesco": "BRADESCO", "Caixa": "Caixa Corporate",
    "BBrasil": "BB", "Bocom": "BOCOM", "ASA": "ASA",
    "BV": "BV", "ABC": "ABC",
}

def importar_bancos(cursor):
    """Garante que todos os bancos existam na tabela bancos."""
    nomes = set(BANCOS_PLANILHA) | set(ENDIVIDAMENTO_PLANILHA.values())
    for nome in sorted(nomes):
        cursor.execute(
            "INSERT IGNORE INTO bancos (nome) VALUES (%s)", (nome,)
        )
    print(f"  ✓ Bancos: {len(nomes)} registros garantidos.")

def banco_id(cursor, nome_planilha):
    """Retorna o id do banco pelo nome (busca parcial)."""
    cursor.execute(
        "SELECT id FROM bancos WHERE nome = %s LIMIT 1", (nome_planilha,)
    )
    r = cursor.fetchone()
    return r[0] if r else None

# ── Aba: Fluxo ────────────────────────────────────────────────────────────────
LINHA_BANCOS      = 1   # linha 1: header com as datas
LINHA_BANCO_INICIO = 2  # linhas 2-13: saldos por banco
LINHA_BANCO_FIM   = 13
LINHA_DUPLICATAS  = 16
LINHA_RECEBIVEIS  = 17
LINHA_HOMEMADE    = 18
LINHA_TOTAL_ENT   = 20
# Saídas
SAIDA_LINHAS = {
    22: "Pagamento a fornecedores / contas de consumo",
    23: "Fornecedor peixe e equipamentos",
    25: "Diversos e empréstimos",
    26: "Câmbio antecipações",
    27: "Câmbios previstos",
    28: "Câmbios efetivamente fechados",
    29: "Desembaraço / fretes",
    30: "Sodexo – cessão Safra",
    31: "Modelez – cessão Safra",
}
LINHA_TOTAL_SAI   = 32
LINHA_SALDO       = 34

BANCO_LINHAS = {
    2: "BRADESCO", 3: "SANTANDER", 4: "ITAU", 5: "BB",
    6: "Safra", 7: "Caixa Corporate", 8: "Bradesco HSBC",
    9: "ABC", 10: "BOCOM", 11: "SICREDI", 12: "ASA", 13: "BV",
}

def importar_fluxo(ws, cursor):
    """Importa a aba Fluxo: saldo_banco, fluxo_entradas, fluxo_saidas, fluxo_saldo."""

    # Lê todas as linhas em memória (somente as primeiras 35 linhas e até coluna 540)
    dados = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=35, max_col=540, values_only=True), 1):
        dados[i] = row

    row1 = dados[1]  # datas nas colunas (índice 0 = col A, 2 = col C, ...)

    # Coleta as colunas de data (col C em diante = índice 2+)
    datas = {}
    for col_idx in range(2, len(row1)):
        v = row1[col_idx]
        d = to_date(v)
        if d:
            datas[col_idx] = d

    if not datas:
        print("  ✗ Nenhuma data encontrada na aba Fluxo.")
        return

    print(f"  ✓ Fluxo: {len(datas)} datas encontradas ({min(datas.values())} a {max(datas.values())})")

    # Categoria_id cache
    cat_ids = {}
    for nome in SAIDA_LINHAS.values():
        cursor.execute("SELECT id FROM categoria_saida WHERE nome = %s", (nome,))
        r = cursor.fetchone()
        if r:
            cat_ids[nome] = r[0]

    registros_saldo   = 0
    registros_banco   = 0
    registros_entradas = 0
    registros_saidas  = 0

    for col_idx, data in datas.items():
        # ── Saldos por banco ─────────────────────────────────────────────
        for linha, nome_banco in BANCO_LINHAS.items():
            row = dados.get(linha, [])
            saldo = to_dec(row[col_idx] if col_idx < len(row) else None)
            bid = banco_id(cursor, nome_banco)
            if bid:
                cursor.execute("""
                    INSERT INTO saldo_banco (banco_id, data, saldo)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE saldo = VALUES(saldo)
                """, (bid, data, saldo))
                registros_banco += 1

        # ── Entradas ─────────────────────────────────────────────────────
        dup  = to_dec(dados.get(LINHA_DUPLICATAS, [])[col_idx] if col_idx < len(dados.get(LINHA_DUPLICATAS, [])) else None)
        rec  = to_dec(dados.get(LINHA_RECEBIVEIS, [])[col_idx] if col_idx < len(dados.get(LINHA_RECEBIVEIS, [])) else None)
        hom  = to_dec(dados.get(LINHA_HOMEMADE, [])[col_idx] if col_idx < len(dados.get(LINHA_HOMEMADE, [])) else None)
        total_ent = dup + rec + hom
        cursor.execute("""
            INSERT INTO fluxo_entradas (data, duplicatas, recebiveis, homemade, total)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
              duplicatas = VALUES(duplicatas),
              recebiveis = VALUES(recebiveis),
              homemade   = VALUES(homemade),
              total      = VALUES(total)
        """, (data, dup, rec, hom, total_ent))
        registros_entradas += 1

        # ── Saídas ───────────────────────────────────────────────────────
        total_sai = Decimal("0")
        for linha, nome_cat in SAIDA_LINHAS.items():
            row = dados.get(linha, [])
            valor = to_dec(row[col_idx] if col_idx < len(row) else None)
            if valor != 0:
                cat_id = cat_ids.get(nome_cat)
                if cat_id:
                    cursor.execute("""
                        INSERT INTO fluxo_saidas (data, categoria_id, valor)
                        VALUES (%s, %s, %s)
                    """, (data, cat_id, valor))
                    registros_saidas += 1
                    total_sai += valor

        # ── Saldo ─────────────────────────────────────────────────────────
        row_saldo = dados.get(LINHA_SALDO, [])
        saldo_val = to_dec(row_saldo[col_idx] if col_idx < len(row_saldo) else None)
        cursor.execute("""
            INSERT INTO fluxo_saldo (data, total_entradas, total_saidas, saldo)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
              total_entradas = VALUES(total_entradas),
              total_saidas   = VALUES(total_saidas),
              saldo          = VALUES(saldo)
        """, (data, total_ent, total_sai, saldo_val))
        registros_saldo += 1

    print(f"  ✓ saldo_banco: {registros_banco} | entradas: {registros_entradas} | saídas: {registros_saidas} | saldo: {registros_saldo}")

# ── Aba: Duplicatas ───────────────────────────────────────────────────────────
def importar_duplicatas(ws, cursor):
    cursor.execute("TRUNCATE TABLE duplicatas")
    total = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        no_titulo = str_limpa(row[0])
        if not no_titulo:
            continue
        cursor.execute("""
            INSERT INTO duplicatas
              (no_titulo, parcela, portador, cod_cliente, nome_cliente,
               dt_emissao, vencimento, vencto_real, vlr_titulo, situacao, banco)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            no_titulo,
            str_limpa(row[1]),
            str_limpa(row[2]),
            str_limpa(row[3]),
            str_limpa(row[4]),
            to_date(row[5]),
            to_date(row[6]),
            to_date(row[7]),
            to_dec(row[8]),
            str_limpa(row[9]),
            str_limpa(row[10]),
        ))
        total += 1
        if total % 50000 == 0:
            print(f"    ... {total} duplicatas inseridas")
    print(f"  ✓ Duplicatas: {total} registros importados.")

# ── Aba: Recebiveis ───────────────────────────────────────────────────────────
def importar_recebiveis(ws, cursor):
    cursor.execute("TRUNCATE TABLE recebiveis")
    total = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        no_titulo = str_limpa(row[0])
        if not no_titulo:
            break  # última linha é total
        cursor.execute("""
            INSERT INTO recebiveis
              (no_titulo, parcela, tipo, portador, cod_cliente, nome_cliente,
               dt_emissao, vencimento, vencto_real, vlr_titulo, desconto_pct, vlr_liquido)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            no_titulo,
            str_limpa(row[1]),
            str_limpa(row[2]),
            str_limpa(row[3]),
            str_limpa(row[4]),
            str_limpa(row[5]),
            to_date(row[6]),
            to_date(row[7]),
            to_date(row[8]),
            to_dec(row[9]),
            to_dec(row[10]),
            to_dec(row[11]),
        ))
        total += 1
    print(f"  ✓ Recebíveis: {total} registros importados.")

# ── Aba: Endividamento ────────────────────────────────────────────────────────
def importar_endividamento(ws, cursor):
    cursor.execute("TRUNCATE TABLE endividamento")
    # Data de referência está na linha 3, col C (índice 2)
    rows = list(ws.iter_rows(min_row=1, max_row=20, max_col=3, values_only=True))
    data_ref = to_date(rows[2][2])  # linha 3 col C
    if not data_ref:
        data_ref = date.today()
    total = 0
    for row in rows[3:]:  # linha 4 em diante
        nome_raw = str_limpa(row[0])
        if not nome_raw or nome_raw.lower().startswith("total"):
            continue
        nome_banco = ENDIVIDAMENTO_PLANILHA.get(nome_raw)
        if not nome_banco:
            continue
        valor = to_dec(row[1])
        bid = banco_id(cursor, nome_banco)
        if bid:
            cursor.execute("""
                INSERT INTO endividamento (banco_id, data_ref, valor_tomado)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE valor_tomado = VALUES(valor_tomado)
            """, (bid, data_ref, valor))
            total += 1
    print(f"  ✓ Endividamento: {total} bancos, data ref {data_ref}.")

# ── Aba: KGIRO ────────────────────────────────────────────────────────────────
KGIRO_BANCOS = ["ITAU", "Santander", "Bradesco", "Safra", "BBrasil", "ABC", "BV", "BOCOM", "SICREDI"]
KGIRO_BANCO_MAP = {
    "ITAU": "ITAU", "SANTANDER": "SANTANDER", "BRADESCO": "BRADESCO",
    "SAFRA": "Safra", "BB": "BB", "ABC": "ABC", "BV": "BV",
    "BOCOM": "BOCOM", "SICREDI": "SICREDI",
}

def importar_kgiro(ws, cursor):
    """Lê contratos de KGIRO: cada bloco começa com banco+n_contrato+data+valor."""
    cursor.execute("DELETE FROM kgiro_parcela")
    cursor.execute("DELETE FROM kgiro_contrato")

    rows = list(ws.iter_rows(min_row=1, max_row=4143, max_col=10, values_only=True))
    contrato_atual = None
    total_contratos = 0
    total_parcelas  = 0

    i = 0
    while i < len(rows):
        row = rows[i]
        banco_nome = str_limpa(row[1]).upper()

        # Detecta início de novo contrato: coluna B tem nome de banco, col C tem número de contrato
        no_contrato = str_limpa(row[2])
        data_c = to_date(row[3])
        valor_c = to_dec(row[4])

        if banco_nome in KGIRO_BANCO_MAP and no_contrato and valor_c > 0 and data_c:
            nome_real = KGIRO_BANCO_MAP[banco_nome]
            bid = banco_id(cursor, nome_real)
            tipo = str_limpa(rows[i+1][2]) if i+1 < len(rows) else ""
            taxa  = str_limpa(rows[i+2][2]) if i+2 < len(rows) else ""
            prazo = 0
            # Conta parcelas: avança até linha vazia ou próximo contrato
            p = i + 1
            while p < len(rows):
                pr = rows[p]
                if str_limpa(pr[1]).upper() in KGIRO_BANCO_MAP and str_limpa(pr[2]) and to_dec(pr[4]) > 0:
                    break
                if to_date(pr[5]):
                    prazo += 1
                p += 1

            cursor.execute("""
                INSERT INTO kgiro_contrato
                  (banco_id, no_contrato, tipo_operacao, taxa, data_contrato, valor_original, prazo_parcelas)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (bid, no_contrato, tipo, taxa, data_c, valor_c, prazo))
            contrato_id = cursor.lastrowid
            contrato_atual = contrato_id
            total_contratos += 1

            # Insere as parcelas deste bloco
            for j in range(i, p):
                pr = rows[j]
                vcto = to_date(pr[5])
                if not vcto:
                    continue
                val_parc = to_dec(pr[6])
                avencer  = to_dec(pr[7])
                principal = to_dec(pr[8])
                pago = 1 if avencer == 0 and val_parc > 0 else 0
                cursor.execute("""
                    INSERT INTO kgiro_parcela
                      (contrato_id, vencimento, valor_parcela, principal, pago)
                    VALUES (%s,%s,%s,%s,%s)
                """, (contrato_id, vcto, val_parc, principal, pago))
                total_parcelas += 1
            i = p
            continue
        i += 1

    print(f"  ✓ KGIRO: {total_contratos} contratos, {total_parcelas} parcelas importadas.")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Uso: python importar_planilha.py caminho/fluxo.xlsx")
        sys.exit(1)

    arquivo = sys.argv[1]
    if not os.path.exists(arquivo):
        print(f"Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    print(f"\n📂 Abrindo planilha: {arquivo}")
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    print(f"   Abas encontradas: {wb.sheetnames}\n")

    print("🔌 Conectando ao banco MySQL...")
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

    try:
        print("\n[1/6] Bancos...")
        importar_bancos(cursor)
        conn.commit()

        print("\n[2/6] Aba Fluxo...")
        importar_fluxo(wb["Fluxo"], cursor)
        conn.commit()

        print("\n[3/6] Aba Duplicatas...")
        importar_duplicatas(wb["Duplicatas"], cursor)
        conn.commit()

        print("\n[4/6] Aba Recebiveis...")
        importar_recebiveis(wb["Recebiveis"], cursor)
        conn.commit()

        print("\n[5/6] Aba KGIRO...")
        importar_kgiro(wb["KGIRO"], cursor)
        conn.commit()

        print("\n[6/6] Aba Endividamento Principal...")
        importar_endividamento(wb["Endividamento Principal "], cursor)
        conn.commit()

        print("\n✅ Importação concluída com sucesso!\n")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ Erro durante importação: {e}")
        raise
    finally:
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        cursor.close()
        conn.close()

if __name__ == "__main__":
    main()
